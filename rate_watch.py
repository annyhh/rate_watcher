import sys
import time
import base64
import traceback

from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import requests
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from openpyxl import Workbook, load_workbook
import ddddocr
import configparser

last_price = None

config = configparser.ConfigParser()
config.read('config.ini', encoding='utf-8')

notify_key = config.get('general', 'notify_key')
interval = config.getint('general', 'interval_seconds')
CURRENCY = config.get('general', 'CURRENCY')
THRESHOLD = 0.5
EXCEL_FILE = config.get('general', 'EXCEL_FILE')
SHEET_NAME = config.get('general', 'SHEET_NAME')

def send_wechat_notify(title, content):

    url = f"https://sctapi.ftqq.com/{notify_key}.send"
    data = {"title": title, "desp": content}
    try:
        resp = requests.post(url, data=data, timeout=10)
        if resp.status_code == 200:
            print("📬 微信通知已发送")
        else:
            print("❌ 通知失败", resp.text)
    except Exception as e:
        print("❌ 发送失败:", str(e))

def write_to_excel(data):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["时间", "币种", "现汇买入价", "现钞买入价", "卖出价"])
    else:
        wb = load_workbook(EXCEL_FILE)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(["时间", "币种", "现汇买入价", "现钞买入价", "卖出价"])
    ws.append(data)
    wb.save(EXCEL_FILE)
    print("✅ 写入 Excel：", data)

def get_code(driver, retry_times=3):
    for attempt in range(retry_times):
        print(f"🔍 第{attempt + 1}次获取验证码...")
        img = driver.find_element(By.ID, "captcha_img")
        img.click()
        time.sleep(1)
        img.screenshot("captcha.png")

        ocr = ddddocr.DdddOcr()
        with open("captcha.png", "rb") as f:
            code = ocr.classification(f.read()).strip()

        print("验证码识别：", repr(code))

        if len(code) == 4 and code.isalnum():
            print("✅ 验证码正确")
            return code
        else:
            print("❌ 验证码错误,重试", code)
    print("❌ 多次获取验证码失败，退出")
    return None

def get_exchange_rate(driver):
    global last_price

    driver.get("https://srh.bankofchina.com/search/whpj/search_cn.jsp")
    Select(driver.find_element(By.NAME, "pjname")).select_by_visible_text(CURRENCY)

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "captcha_img"))
    )

    code = get_code(driver)

    driver.find_element(By.NAME, "captcha").send_keys(code)
    driver.find_element(By.XPATH, '//input[@value="查询"]').click()

    time.sleep(2)
    rows = driver.find_elements(By.XPATH, '//table[@align="left"]//tr')[1:]

    if not rows:
        print("❌ 查询失败，可能验证码错误")
        return

    cols = rows[0].find_elements(By.TAG_NAME, "td")
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    xianhui = None
    if len(cols) >= 4:
        xianhui = float(cols[1].text.strip())
        data = [now, CURRENCY, xianhui, cols[2].text.strip(), cols[3].text.strip()]
        write_to_excel(data)
    else:
        print("⚠️ 跳过：列数不足，内容：", [c.text for c in cols])

    if last_price is not None and xianhui is not None:
        delta = round(abs(xianhui - last_price), 4)
        if delta >= THRESHOLD:
            send_wechat_notify(
                f"当前现汇买入价：{xianhui}\n上次：{last_price}\n波动：{delta}",
                f"💱 美元汇率变动：{last_price} ➜ {xianhui}（{delta:+.2f}）"
            )
    last_price = xianhui

def resource_path(relative_path):
    import sys
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.abspath(relative_path)

def main():
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_path = resource_path("chromedriver.exe")
    service = Service(chrome_path)
    driver = webdriver.Chrome(service=service, options=options)

    print(f"🌀 汇率监控启动中（币种：{CURRENCY}）...")

    try:
        while True:
            get_exchange_rate(driver)
            print(f"⏳ 等待 {interval // 60} 分钟...\n")
            time.sleep(interval)
    except KeyboardInterrupt:
        driver.quit()
        print("🛑 已退出")
    except Exception as e:
        print("❌ 出错了：", str(e))
        traceback.print_exc()
        driver.quit()


if __name__ == "__main__":
    main()
